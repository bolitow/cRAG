# src/ingestion/parsers/excel_parser.py
import os
import re
from typing import List, Dict, Optional, Tuple, Any
from datetime import datetime
import logging

# Imports pour le traitement Excel
try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd

    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

from .base_parser import BaseParser, ParsedDocument, ParsingError


class ExcelParser(BaseParser):
    """
    Parser spécialisé pour les fichiers Excel (.xlsx, .xls).

    Ce parser est conçu pour extraire intelligemment les données des feuilles Excel,
    en comprenant les structures courantes dans les documents de cybersécurité :
    - Matrices de contrôle (contrôles vs exigences)
    - Registres de risques
    - Inventaires d'actifs
    - Tableaux de conformité
    - Listes de configurations

    Le parser identifie automatiquement la structure des données et les convertit
    en format textuel enrichi pour le traitement par le système CRAG.
    """

    def __init__(self, config: Optional[Dict] = None):
        """
        Initialise le parser Excel.

        Config options:
            - sheet_names: Liste des noms de feuilles à traiter (None = toutes)
            - header_row: Numéro de ligne des en-têtes (auto-détection par défaut)
            - skip_empty_sheets: Ignorer les feuilles vides
            - merge_cells_handling: Comment gérer les cellules fusionnées
            - detect_tables: Détecter automatiquement les zones de tableau
            - preserve_formulas: Garder les formules ou juste les valeurs
        """
        super().__init__(config)
        self.supported_extensions = ['.xlsx', '.xls', '.xlsm', '.xlsb']

        # Vérifier les bibliothèques disponibles
        if not any([PANDAS_AVAILABLE, OPENPYXL_AVAILABLE, XLRD_AVAILABLE]):
            raise ImportError(
                "Aucune bibliothèque Excel trouvée. Installez au moins une de ces options:\n"
                "pip install pandas openpyxl xlrd"
            )

        # Configuration
        self.sheet_names = config.get('sheet_names') if config else None
        self.header_row = config.get('header_row') if config else None
        self.skip_empty_sheets = config.get('skip_empty_sheets', True) if config else True
        self.merge_cells_handling = config.get('merge_cells_handling', 'expand') if config else 'expand'
        self.detect_tables = config.get('detect_tables', True) if config else True
        self.preserve_formulas = config.get('preserve_formulas', False) if config else False

        # Logger
        self.logger = logging.getLogger(__name__)

        # Patterns pour identifier les types de contenu
        self._init_content_patterns()

    def _init_content_patterns(self):
        """
        Initialise les patterns pour identifier les types de tableaux.

        Ces patterns aident à comprendre le contexte et la structure
        des données Excel pour un meilleur traitement.
        """
        # Patterns pour les matrices de contrôle
        self.control_matrix_headers = [
            r'contr[ôo]le', r'control', r'mesure', r'measure',
            r'exigence', r'requirement', r'r[éè]f[ée]rence', r'reference',
            r'statut', r'status', r'[ée]tat', r'state',
            r'responsable', r'owner', r'criticité', r'criticality',
            r'conforme', r'compliant', r'conformit[ée]', r'compliance'
        ]

        # Patterns pour les registres de risques
        self.risk_register_headers = [
            r'risque', r'risk', r'menace', r'threat',
            r'impact', r'probabilit[ée]', r'probability',
            r'niveau', r'level', r'score',
            r'mitigation', r'traitement', r'treatment'
        ]

        # Patterns pour les inventaires
        self.inventory_headers = [
            r'actif', r'asset', r'nom', r'name',
            r'type', r'cat[ée]gorie', r'category',
            r'propri[ée]taire', r'owner', r'localisation', r'location',
            r'version', r'criticité', r'criticality'
        ]

        # Patterns pour identifier les statuts
        self.status_patterns = {
            'compliant': [r'conforme', r'ok', r'compliant', r'pass', r'vert', r'green'],
            'non_compliant': [r'non.?conforme', r'nok', r'ko', r'fail', r'rouge', r'red'],
            'partial': [r'partiel', r'partial', r'orange', r'jaune', r'yellow'],
            'na': [r'n/?a', r'non.?applicable', r'not.?applicable']
        }

    def can_parse(self, file_path: str) -> bool:
        """
        Vérifie si le fichier est un Excel valide.
        """
        if not os.path.exists(file_path):
            return False

        # Vérifier l'extension
        ext = os.path.splitext(file_path)[1].lower()
        return ext in self.supported_extensions

    def parse(self, file_path: str) -> ParsedDocument:
        """
        Parse le fichier Excel et extrait toutes les données structurées.

        Cette méthode identifie intelligemment le type de données
        et les convertit en format textuel enrichi.
        """
        if not self.can_parse(file_path):
            raise ParsingError(f"Le fichier {file_path} n'est pas un fichier Excel valide")

        self.logger.info(f"Parsing du fichier Excel : {file_path}")

        # Initialiser les variables
        metadata = self.extract_metadata(file_path)
        content_parts = []
        tables = []
        parsing_errors = []
        sections = []

        try:
            # Lire le fichier Excel
            if PANDAS_AVAILABLE:
                excel_data = self._read_with_pandas(file_path)
            elif OPENPYXL_AVAILABLE:
                excel_data = self._read_with_openpyxl(file_path)
            else:
                excel_data = self._read_with_xlrd(file_path)

            # Traiter chaque feuille
            for sheet_name, sheet_data in excel_data.items():
                if self.skip_empty_sheets and self._is_sheet_empty(sheet_data):
                    continue

                # Analyser le type de contenu
                content_type = self._identify_content_type(sheet_data)

                # Convertir en format textuel selon le type
                if content_type == "control_matrix":
                    sheet_content, sheet_tables = self._process_control_matrix(sheet_name, sheet_data)
                elif content_type == "risk_register":
                    sheet_content, sheet_tables = self._process_risk_register(sheet_name, sheet_data)
                elif content_type == "inventory":
                    sheet_content, sheet_tables = self._process_inventory(sheet_name, sheet_data)
                else:
                    sheet_content, sheet_tables = self._process_generic_table(sheet_name, sheet_data)

                # Ajouter au contenu global
                content_parts.append(sheet_content)
                tables.extend(sheet_tables)

                # Créer une section pour cette feuille
                sections.append({
                    "title": f"Feuille : {sheet_name}",
                    "level": 1,
                    "content": sheet_content,
                    "type": content_type
                })

            # Assembler le contenu final
            content = "\n\n".join(content_parts)

            # Ajouter les métadonnées Excel
            metadata.update({
                'format': 'excel',
                'sheets_count': len(excel_data),
                'sheets_processed': len(sections),
                'total_tables': len(tables)
            })

            # Identifier le type de document
            doc_type = self._identify_excel_doc_type(sections, metadata)

            # Calculer le hash et la confiance
            doc_hash = self.calculate_document_hash(content)
            confidence = self._calculate_confidence_score(content, tables, sections, parsing_errors)

        except Exception as e:
            self.logger.error(f"Erreur lors du parsing : {str(e)}")
            parsing_errors.append(f"Erreur : {str(e)}")
            content = ""
            confidence = 0.0
            doc_type = "excel"
            doc_hash = ""

        # Créer le document parsé
        parsed_doc = ParsedDocument(
            content=content,
            metadata=metadata,
            source_path=file_path,
            doc_type=doc_type,
            sections=sections,
            tables=tables,
            images_count=0,  # Excel n'a pas vraiment d'images inline
            parsing_errors=parsing_errors,
            parsing_timestamp=datetime.now(),
            document_hash=doc_hash,
            confidence_score=confidence
        )

        # Valider et mettre à jour les stats
        is_valid, warnings = self.validate_parsing_result(parsed_doc)
        self.update_stats(is_valid, len(warnings))

        return parsed_doc

    def _read_with_pandas(self, file_path: str) -> Dict[str, Any]:
        """
        Lit le fichier Excel avec pandas.

        Pandas est excellent pour gérer les données tabulaires complexes
        et offre de nombreuses options de lecture.
        """
        excel_data = {}

        try:
            # Lire toutes les feuilles
            excel_file = pd.ExcelFile(file_path)

            # Filtrer les feuilles si nécessaire
            sheets_to_read = self.sheet_names or excel_file.sheet_names

            for sheet_name in sheets_to_read:
                if sheet_name in excel_file.sheet_names:
                    # Lire la feuille avec gestion intelligente des headers
                    df = pd.read_excel(
                        excel_file,
                        sheet_name=sheet_name,
                        header=self.header_row
                    )

                    # Convertir en structure de données standard
                    excel_data[sheet_name] = {
                        'data': df.values.tolist(),
                        'columns': df.columns.tolist(),
                        'shape': df.shape,
                        'df': df  # Garder le DataFrame pour des analyses avancées
                    }

        except Exception as e:
            raise ParsingError(f"Erreur pandas : {str(e)}")

        return excel_data

    def _read_with_openpyxl(self, file_path: str) -> Dict[str, Any]:
        """
        Lit le fichier Excel avec openpyxl.

        Openpyxl permet un accès de bas niveau aux cellules,
        utile pour préserver les formules et la mise en forme.
        """
        from openpyxl import load_workbook

        excel_data = {}

        try:
            # Charger le workbook
            wb = load_workbook(
                file_path,
                data_only=not self.preserve_formulas
            )

            # Filtrer les feuilles
            sheets_to_read = self.sheet_names or wb.sheetnames

            for sheet_name in sheets_to_read:
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]

                    # Extraire les données
                    data = []
                    for row in sheet.iter_rows(values_only=True):
                        data.append(list(row))

                    # Détecter les colonnes (première ligne non vide)
                    columns = []
                    for row in data:
                        if any(cell is not None for cell in row):
                            columns = [str(cell) if cell is not None else f"Col{i + 1}"
                                       for i, cell in enumerate(row)]
                            break

                    excel_data[sheet_name] = {
                        'data': data,
                        'columns': columns,
                        'shape': (sheet.max_row, sheet.max_column),
                        'merged_cells': [str(cell) for cell in sheet.merged_cells.ranges]
                    }

            wb.close()

        except Exception as e:
            raise ParsingError(f"Erreur openpyxl : {str(e)}")

        return excel_data

    def _read_with_xlrd(self, file_path: str) -> Dict[str, Any]:
        """
        Lit le fichier Excel avec xlrd (pour les anciens formats .xls).
        """
        import xlrd

        excel_data = {}

        try:
            workbook = xlrd.open_workbook(file_path)

            # Filtrer les feuilles
            sheets_to_read = self.sheet_names or workbook.sheet_names()

            for sheet_name in sheets_to_read:
                if sheet_name in workbook.sheet_names():
                    sheet = workbook.sheet_by_name(sheet_name)

                    # Extraire les données
                    data = []
                    for row_idx in range(sheet.nrows):
                        row_data = []
                        for col_idx in range(sheet.ncols):
                            cell = sheet.cell(row_idx, col_idx)
                            row_data.append(cell.value)
                        data.append(row_data)

                    # Détecter les colonnes
                    columns = data[0] if data else []

                    excel_data[sheet_name] = {
                        'data': data,
                        'columns': columns,
                        'shape': (sheet.nrows, sheet.ncols)
                    }

        except Exception as e:
            raise ParsingError(f"Erreur xlrd : {str(e)}")

        return excel_data

    def _is_sheet_empty(self, sheet_data: Dict) -> bool:
        """
        Vérifie si une feuille est vide ou ne contient que des cellules vides.
        """
        if not sheet_data.get('data'):
            return True

        # Vérifier si toutes les cellules sont vides
        for row in sheet_data['data']:
            if any(cell is not None and str(cell).strip() for cell in row):
                return False

        return True

    def _identify_content_type(self, sheet_data: Dict) -> str:
        """
        Identifie le type de contenu dans la feuille Excel.

        Cette identification permet d'appliquer le traitement approprié
        selon que c'est une matrice de contrôle, un registre de risques, etc.
        """
        # Obtenir les headers (première ligne non vide avec du contenu)
        headers = []
        for row in sheet_data.get('data', []):
            if any(cell for cell in row if cell):
                headers = [str(cell).lower() if cell else '' for cell in row]
                break

        if not headers:
            return "generic"

        # Compter les correspondances pour chaque type
        scores = {
            'control_matrix': 0,
            'risk_register': 0,
            'inventory': 0
        }

        # Vérifier les patterns de contrôle
        for header in headers:
            for pattern in self.control_matrix_headers:
                if re.search(pattern, header, re.IGNORECASE):
                    scores['control_matrix'] += 1
                    break

        # Vérifier les patterns de risque
        for header in headers:
            for pattern in self.risk_register_headers:
                if re.search(pattern, header, re.IGNORECASE):
                    scores['risk_register'] += 1
                    break

        # Vérifier les patterns d'inventaire
        for header in headers:
            for pattern in self.inventory_headers:
                if re.search(pattern, header, re.IGNORECASE):
                    scores['inventory'] += 1
                    break

        # Retourner le type avec le score le plus élevé
        max_score = max(scores.values())
        if max_score >= 2:  # Au moins 2 correspondances
            return max(scores, key=scores.get)

        return "generic"

    def _process_control_matrix(self, sheet_name: str, sheet_data: Dict) -> Tuple[str, List[Dict]]:
        """
        Traite une matrice de contrôle de sécurité.

        Les matrices de contrôle ont généralement une structure :
        - Lignes : Contrôles ou exigences
        - Colonnes : Attributs (statut, responsable, description, etc.)

        Cette méthode extrait et structure ces informations pour le système CRAG.
        """
        content_parts = [f"# Matrice de Contrôle : {sheet_name}\n"]
        tables = []

        data = sheet_data.get('data', [])
        if not data:
            return "", []

        # Identifier les headers
        headers = data[0] if data else []

        # Identifier les colonnes clés
        control_col = self._find_column_index(headers, ['contrôle', 'control', 'mesure', 'id'])
        desc_col = self._find_column_index(headers, ['description', 'desc', 'détail'])
        status_col = self._find_column_index(headers, ['statut', 'status', 'état', 'conformité'])
        owner_col = self._find_column_index(headers, ['responsable', 'owner', 'propriétaire'])
        ref_col = self._find_column_index(headers, ['référence', 'reference', 'norme', 'standard'])

        # Statistiques
        stats = {
            'total': 0,
            'conforme': 0,
            'non_conforme': 0,
            'partiel': 0,
            'na': 0
        }

        # Grouper par référence/norme si disponible
        controls_by_ref = {}

        # Traiter chaque ligne de contrôle
        for row_idx, row in enumerate(data[1:], 1):  # Skip header
            if not any(cell for cell in row if cell):  # Skip empty rows
                continue

            stats['total'] += 1

            # Extraire les informations
            control_id = row[control_col] if control_col is not None else f"CTRL-{row_idx}"
            description = row[desc_col] if desc_col is not None else "Description non disponible"
            status = row[status_col] if status_col is not None else "Non défini"
            owner = row[owner_col] if owner_col is not None else "Non assigné"
            reference = row[ref_col] if ref_col is not None else "Général"

            # Classifier le statut
            status_lower = str(status).lower()
            status_type = self._classify_status(status_lower)
            stats[status_type] += 1

            # Grouper par référence
            ref_key = str(reference)
            if ref_key not in controls_by_ref:
                controls_by_ref[ref_key] = []

            controls_by_ref[ref_key].append({
                'id': control_id,
                'description': description,
                'status': status,
                'status_type': status_type,
                'owner': owner
            })

        # Générer le contenu textuel structuré
        content_parts.append(f"## Résumé de Conformité\n")
        content_parts.append(f"- Total des contrôles : {stats['total']}")
        content_parts.append(f"- Conformes : {stats['conforme']} ({stats['conforme'] / stats['total'] * 100:.1f}%)")
        content_parts.append(
            f"- Non conformes : {stats['non_conforme']} ({stats['non_conforme'] / stats['total'] * 100:.1f}%)")
        content_parts.append(f"- Partiellement conformes : {stats['partiel']}")
        content_parts.append(f"- Non applicables : {stats['na']}\n")

        # Détailler par référence/norme
        for ref, controls in controls_by_ref.items():
            content_parts.append(f"\n## Contrôles - {ref}\n")

            # Séparer par statut
            non_conformes = [c for c in controls if c['status_type'] == 'non_conforme']
            if non_conformes:
                content_parts.append("### ⚠️ Contrôles Non Conformes\n")
                for ctrl in non_conformes:
                    content_parts.append(
                        f"- **{ctrl['id']}** : {ctrl['description']}\n"
                        f"  - Statut : {ctrl['status']}\n"
                        f"  - Responsable : {ctrl['owner']}\n"
                    )

            partiels = [c for c in controls if c['status_type'] == 'partiel']
            if partiels:
                content_parts.append("\n### ⚡ Contrôles Partiellement Conformes\n")
                for ctrl in partiels:
                    content_parts.append(
                        f"- **{ctrl['id']}** : {ctrl['description']}\n"
                        f"  - Statut : {ctrl['status']}\n"
                        f"  - Responsable : {ctrl['owner']}\n"
                    )

            conformes = [c for c in controls if c['status_type'] == 'conforme']
            if conformes:
                content_parts.append(f"\n### ✅ Contrôles Conformes : {len(conformes)} contrôles\n")

        # Créer un tableau récapitulatif
        table_data = [headers]
        table_data.extend(data[1:stats['total'] + 1])

        tables.append({
            "type": "control_matrix",
            "title": f"Matrice de Contrôle - {sheet_name}",
            "data": table_data,
            "stats": stats,
            "text_representation": self._table_to_text(table_data, "control_matrix")
        })

        return "\n".join(content_parts), tables

    def _process_risk_register(self, sheet_name: str, sheet_data: Dict) -> Tuple[str, List[Dict]]:
        """
        Traite un registre de risques.

        Les registres de risques contiennent typiquement :
        - Identification du risque
        - Évaluation (impact, probabilité)
        - Mesures de mitigation
        - Responsable
        """
        content_parts = [f"# Registre des Risques : {sheet_name}\n"]
        tables = []

        data = sheet_data.get('data', [])
        if not data:
            return "", []

        headers = data[0] if data else []

        # Identifier les colonnes clés
        risk_col = self._find_column_index(headers, ['risque', 'risk', 'menace', 'threat'])
        impact_col = self._find_column_index(headers, ['impact'])
        prob_col = self._find_column_index(headers, ['probabilité', 'probability', 'prob'])
        level_col = self._find_column_index(headers, ['niveau', 'level', 'criticité'])
        mitigation_col = self._find_column_index(headers, ['mitigation', 'traitement', 'mesure'])

        # Classifier les risques par niveau
        risks_by_level = {
            'critique': [],
            'élevé': [],
            'moyen': [],
            'faible': []
        }

        # Traiter chaque risque
        for row_idx, row in enumerate(data[1:], 1):
            if not any(cell for cell in row if cell):
                continue

            risk_desc = row[risk_col] if risk_col is not None else f"Risque {row_idx}"
            impact = row[impact_col] if impact_col is not None else "Non évalué"
            probability = row[prob_col] if prob_col is not None else "Non évalué"
            level = row[level_col] if level_col is not None else "Non défini"
            mitigation = row[mitigation_col] if mitigation_col is not None else "Aucune"

            # Classifier le niveau
            level_class = self._classify_risk_level(str(level).lower())

            risk_info = {
                'description': risk_desc,
                'impact': impact,
                'probability': probability,
                'level': level,
                'mitigation': mitigation
            }

            risks_by_level[level_class].append(risk_info)

        # Générer le contenu structuré
        total_risks = sum(len(risks) for risks in risks_by_level.values())
        content_parts.append(f"## Résumé des Risques\n")
        content_parts.append(f"- Total des risques identifiés : {total_risks}")
        content_parts.append(f"- Risques critiques : {len(risks_by_level['critique'])}")
        content_parts.append(f"- Risques élevés : {len(risks_by_level['élevé'])}")
        content_parts.append(f"- Risques moyens : {len(risks_by_level['moyen'])}")
        content_parts.append(f"- Risques faibles : {len(risks_by_level['faible'])}\n")

        # Détailler par niveau de criticité
        for level, risks in risks_by_level.items():
            if risks:
                emoji = {'critique': '🔴', 'élevé': '🟠', 'moyen': '🟡', 'faible': '🟢'}
                content_parts.append(f"\n## {emoji[level]} Risques {level.capitalize()}\n")

                for risk in risks:
                    content_parts.append(f"### {risk['description']}\n")
                    content_parts.append(f"- **Impact** : {risk['impact']}")
                    content_parts.append(f"- **Probabilité** : {risk['probability']}")
                    content_parts.append(f"- **Niveau** : {risk['level']}")
                    content_parts.append(f"- **Mitigation** : {risk['mitigation']}\n")

        # Créer le tableau
        tables.append({
            "type": "risk_register",
            "title": f"Registre des Risques - {sheet_name}",
            "data": data[:total_risks + 1],
            "summary": {level: len(risks) for level, risks in risks_by_level.items()},
            "text_representation": self._table_to_text(data[:total_risks + 1], "risk_register")
        })

        return "\n".join(content_parts), tables

    def _process_inventory(self, sheet_name: str, sheet_data: Dict) -> Tuple[str, List[Dict]]:
        """
        Traite un inventaire d'actifs.

        Les inventaires contiennent typiquement :
        - Identification de l'actif
        - Type/Catégorie
        - Propriétaire
        - Criticité
        - Localisation
        """
        content_parts = [f"# Inventaire des Actifs : {sheet_name}\n"]
        tables = []

        data = sheet_data.get('data', [])
        if not data:
            return "", []

        headers = data[0] if data else []

        # Identifier les colonnes
        name_col = self._find_column_index(headers, ['nom', 'name', 'actif', 'asset', 'id'])
        type_col = self._find_column_index(headers, ['type', 'catégorie', 'category'])
        owner_col = self._find_column_index(headers, ['propriétaire', 'owner', 'responsable'])
        criticality_col = self._find_column_index(headers, ['criticité', 'criticality', 'importance'])

        # Grouper par type et criticité
        assets_by_type = {}
        criticality_stats = {'critique': 0, 'élevée': 0, 'moyenne': 0, 'faible': 0}

        total_assets = 0
        for row in data[1:]:
            if not any(cell for cell in row if cell):
                continue

            total_assets += 1

            asset_name = row[name_col] if name_col is not None else f"Actif {total_assets}"
            asset_type = row[type_col] if type_col is not None else "Non catégorisé"
            owner = row[owner_col] if owner_col is not None else "Non assigné"
            criticality = row[criticality_col] if criticality_col is not None else "Non définie"

            # Classifier la criticité
            crit_class = self._classify_criticality(str(criticality).lower())
            criticality_stats[crit_class] += 1

            # Grouper par type
            if asset_type not in assets_by_type:
                assets_by_type[asset_type] = []

            assets_by_type[asset_type].append({
                'name': asset_name,
                'owner': owner,
                'criticality': criticality,
                'criticality_class': crit_class
            })

        # Générer le contenu
        content_parts.append(f"## Résumé de l'Inventaire\n")
        content_parts.append(f"- Total des actifs : {total_assets}")
        content_parts.append(f"- Types d'actifs : {len(assets_by_type)}")
        content_parts.append(f"\n### Distribution par Criticité")
        content_parts.append(f"- Critiques : {criticality_stats['critique']}")
        content_parts.append(f"- Élevée : {criticality_stats['élevée']}")
        content_parts.append(f"- Moyenne : {criticality_stats['moyenne']}")
        content_parts.append(f"- Faible : {criticality_stats['faible']}\n")

        # Détailler par type
        for asset_type, assets in assets_by_type.items():
            content_parts.append(f"\n## {asset_type} ({len(assets)} actifs)\n")

            # Séparer par criticité
            critical_assets = [a for a in assets if a['criticality_class'] == 'critique']
            if critical_assets:
                content_parts.append("### 🔴 Actifs Critiques\n")
                for asset in critical_assets:
                    content_parts.append(f"- **{asset['name']}** - Propriétaire : {asset['owner']}")

            other_assets = [a for a in assets if a['criticality_class'] != 'critique']
            if other_assets:
                content_parts.append(f"\n### Autres Actifs ({len(other_assets)})")
                # Lister simplement les autres
                for asset in other_assets[:5]:  # Limiter pour ne pas surcharger
                    content_parts.append(f"- {asset['name']} ({asset['criticality']})")
                if len(other_assets) > 5:
                    content_parts.append(f"- ... et {len(other_assets) - 5} autres")

        tables.append({
            "type": "inventory",
            "title": f"Inventaire - {sheet_name}",
            "data": data[:total_assets + 1],
            "stats": {
                "total": total_assets,
                "by_type": {t: len(a) for t, a in assets_by_type.items()},
                "by_criticality": criticality_stats
            },
            "text_representation": self._table_to_text(data[:total_assets + 1], "inventory")
        })

        return "\n".join(content_parts), tables

    def _process_generic_table(self, sheet_name: str, sheet_data: Dict) -> Tuple[str, List[Dict]]:
        """
        Traite un tableau générique sans structure spécifique identifiée.

        Cette méthode fait de son mieux pour extraire l'information
        de manière structurée même sans connaître le schéma exact.
        """
        content_parts = [f"# Tableau : {sheet_name}\n"]
        tables = []

        data = sheet_data.get('data', [])
        if not data:
            return "", []

        # Trouver la première ligne non vide pour les headers
        headers = []
        data_start_idx = 0
        for idx, row in enumerate(data):
            if any(cell for cell in row if cell):
                headers = [str(cell) if cell else f"Col{i + 1}" for i, cell in enumerate(row)]
                data_start_idx = idx + 1
                break

        # Compter les lignes de données non vides
        data_rows = []
        for row in data[data_start_idx:]:
            if any(cell for cell in row if cell):
                data_rows.append(row)

        content_parts.append(f"## Informations du Tableau\n")
        content_parts.append(f"- Colonnes : {len(headers)}")
        content_parts.append(f"- Lignes de données : {len(data_rows)}\n")

        # Afficher les headers
        content_parts.append("## Structure des Colonnes\n")
        for i, header in enumerate(headers):
            if header and not header.startswith("Col"):
                content_parts.append(f"- Colonne {i + 1} : {header}")

        # Échantillon de données
        if data_rows:
            content_parts.append("\n## Échantillon de Données\n")
            sample_size = min(5, len(data_rows))

            for row_idx, row in enumerate(data_rows[:sample_size]):
                content_parts.append(f"\n### Ligne {row_idx + 1}")
                for col_idx, (header, value) in enumerate(zip(headers, row)):
                    if value is not None and str(value).strip():
                        content_parts.append(f"- **{header}** : {value}")

            if len(data_rows) > sample_size:
                content_parts.append(f"\n... et {len(data_rows) - sample_size} autres lignes")

        # Créer le tableau complet pour référence
        full_data = [headers] + data_rows
        tables.append({
            "type": "generic",
            "title": sheet_name,
            "data": full_data,
            "shape": (len(full_data), len(headers)),
            "text_representation": self._table_to_text(full_data, "generic")
        })

        return "\n".join(content_parts), tables

    def _find_column_index(self, headers: List, keywords: List[str]) -> Optional[int]:
        """
        Trouve l'index d'une colonne basé sur des mots-clés.

        Cette méthode permet de gérer les variations dans les noms de colonnes.
        """
        for idx, header in enumerate(headers):
            if header is None:
                continue

            header_lower = str(header).lower()
            for keyword in keywords:
                if keyword.lower() in header_lower:
                    return idx

        return None

    def _classify_status(self, status: str) -> str:
        """
        Classifie un statut de conformité.
        """
        for status_type, patterns in self.status_patterns.items():
            for pattern in patterns:
                if re.search(pattern, status, re.IGNORECASE):
                    return status_type.replace('_', '_')

        return 'non_défini'

    def _classify_risk_level(self, level: str) -> str:
        """
        Classifie un niveau de risque.
        """
        if any(word in level for word in ['critique', 'critical', 'très élevé', 'very high']):
            return 'critique'
        elif any(word in level for word in ['élevé', 'high', 'important']):
            return 'élevé'
        elif any(word in level for word in ['moyen', 'medium', 'modéré']):
            return 'moyen'
        elif any(word in level for word in ['faible', 'low', 'mineur']):
            return 'faible'
        else:
            return 'moyen'  # Par défaut

    def _classify_criticality(self, criticality: str) -> str:
        """
        Classifie un niveau de criticité d'actif.
        """
        if any(word in criticality for word in ['critique', 'critical', 'vital']):
            return 'critique'
        elif any(word in criticality for word in ['élevé', 'high', 'important']):
            return 'élevée'
        elif any(word in criticality for word in ['moyen', 'medium', 'normal']):
            return 'moyenne'
        else:
            return 'faible'

    def _identify_excel_doc_type(self, sections: List[Dict], metadata: Dict) -> str:
        """
        Identifie le type de document Excel basé sur son contenu.
        """
        # Compter les types de sections
        section_types = [s.get('type', 'generic') for s in sections]

        if 'control_matrix' in section_types:
            return 'control_matrix'
        elif 'risk_register' in section_types:
            return 'risk_assessment'
        elif 'inventory' in section_types:
            return 'asset_inventory'
        else:
            # Vérifier le nom du fichier
            filename = metadata.get('file_name', '').lower()
            if 'control' in filename or 'matrice' in filename:
                return 'control_matrix'
            elif 'risk' in filename or 'risque' in filename:
                return 'risk_assessment'
            elif 'inventory' in filename or 'inventaire' in filename:
                return 'asset_inventory'

        return 'data_table'

    def _table_to_text(self, table_data: List[List], table_type: str) -> str:
        """
        Convertit un tableau en texte structuré adapté au type.

        Cette méthode produit un format optimisé pour le traitement
        par le système CRAG.
        """
        if not table_data or len(table_data) < 2:
            return ""

        headers = table_data[0]
        rows = table_data[1:]

        text_parts = []

        if table_type == "control_matrix":
            text_parts.append("Matrice de Contrôle de Sécurité")
            text_parts.append("=" * 50)

            # Format spécial pour les contrôles
            for row in rows:
                if not any(cell for cell in row if cell):
                    continue

                control_text = []
                for i, (header, value) in enumerate(zip(headers, row)):
                    if value and str(value).strip():
                        if 'contrôle' in str(header).lower() or 'id' in str(header).lower():
                            control_text.insert(0, f"Contrôle {value}:")
                        elif 'statut' in str(header).lower():
                            control_text.append(f"[{value}]")
                        else:
                            control_text.append(f"{header}: {value}")

                text_parts.append(" | ".join(control_text))

        elif table_type == "risk_register":
            text_parts.append("Registre des Risques")
            text_parts.append("=" * 50)

            for row in rows:
                if not any(cell for cell in row if cell):
                    continue

                risk_parts = []
                for header, value in zip(headers, row):
                    if value and str(value).strip():
                        risk_parts.append(f"{header}: {value}")

                text_parts.append("Risque: " + " | ".join(risk_parts))

        else:
            # Format générique
            text_parts.append("Tableau de Données")
            text_parts.append("Colonnes: " + " | ".join(str(h) for h in headers))
            text_parts.append("-" * 50)

            for i, row in enumerate(rows[:10]):  # Limiter pour éviter trop de texte
                if any(cell for cell in row if cell):
                    row_text = f"L{i + 1}: " + " | ".join(
                        str(cell) if cell else "-" for cell in row
                    )
                    text_parts.append(row_text)

            if len(rows) > 10:
                text_parts.append(f"... et {len(rows) - 10} autres lignes")

        return "\n".join(text_parts)

    def _calculate_confidence_score(self, content: str, tables: List[Dict],
                                    sections: List[Dict], errors: List[str]) -> float:
        """
        Calcule le score de confiance pour le parsing Excel.
        """
        score = 1.0

        # Pénalités pour les erreurs
        score -= len(errors) * 0.2

        # Vérifier qu'on a extrait du contenu
        if not content or len(content) < 50:
            score -= 0.5

        # Bonus si on a identifié le type de contenu
        identified_types = [s.get('type') for s in sections if s.get('type') != 'generic']
        if identified_types:
            score += 0.1

        # Bonus si on a des tables structurées
        if tables:
            score += min(0.2, len(tables) * 0.05)

        # Vérifier la cohérence des données
        for table in tables:
            if 'stats' in table or 'summary' in table:
                score += 0.05  # Bonus pour l'analyse statistique

        return max(0.0, min(1.0, score))